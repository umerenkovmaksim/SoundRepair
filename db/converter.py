from openpyxl import load_workbook
import sqlite3

con = sqlite3.connect("data.db")
cur = con.cursor()
id_list = list(map(lambda x: x[0], cur.execute("""SELECT id from products""")))
con.close()
print(id_list)

wb = load_workbook('data.xlsx')
sheet = wb['Лист1']

for cellObj in sheet['2':f'{sheet.max_row}']:
    product_data = list(map(lambda x: x.value, list(cellObj)))
    print(product_data)
    if product_data[0] in id_list:
        con = sqlite3.connect("data.db")
        cur = con.cursor()
        cur.execute(f"""UPDATE products
                        SET
                        categorie = '{product_data[1]}', 
                        subcategorie = '{product_data[2]}', 
                        manufacturer = '{product_data[3]}', 
                        name = '{product_data[4]}', 
                        description = '{product_data[5]}',  
                        quantity = {product_data[6]}, 
                        price = {product_data[7]},
                        sale = 0
                        WHERE id == {product_data[0]}""")
        con.commit()
        con.close()

    else:
        product_data.append(0)

        con = sqlite3.connect("data.db")
        cur = con.cursor()
        cur.execute(
            f"""INSERT INTO 
            products(id, categorie, subcategorie, manufacturer, name, description, quantity, price, sale) 
            VALUES
            {tuple(product_data)}""")

        con.commit()
        con.close()
